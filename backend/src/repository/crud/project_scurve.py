import datetime
import decimal
import pathlib
import typing
import uuid
from io import BytesIO

import openpyxl
import sqlalchemy

from src.models.db.project_progress import ProjectProgress
from src.models.db.project_progress_summary import ProjectProgressSummary
from src.models.db.project_scurve_upload import ProjectScurveUpload
from src.repository.crud.base import BaseCRUDRepository

# ponytail: global lock-free file storage on local disk — fine for this app's single-instance
# deployment; move to object storage if it's ever run behind multiple backend replicas.
UPLOAD_ROOT = pathlib.Path(__file__).resolve().parents[3] / "uploads" / "scurve"

SUMMARY_HEADER_ROW = 58
SUMMARY_DATE_ROW = 59
SUMMARY_METRIC_ROWS = {
    "actual_this_week": 64,
    "actual_cumulative": 65,
    "plan_this_week": 66,
    "plan_cumulative": 67,
    "variance_to_plan": 70,
}
SUMMARY_FIRST_DATA_COL = 3  # column C

PROGRESS_HEADER_ROW = 19  # "No" / "Description" / "WF" / ...
PROGRESS_FIRST_DATA_ROW = 22
PROGRESS_MAX_ROWS_SCANNED = 200  # safety cap — stop scanning if "Overall" is never found
PROGRESS_COLUMNS = {
    "item_no": 2,  # B
    "description": 3,  # C
    "wf": 4,  # D
    "previous_week_plan": 5,  # E
    "previous_week_actual": 7,  # G
    "previous_week_variance": 8,  # H
    "this_week_plan": 10,  # J
    "this_week_actual": 12,  # L
    "this_week_variance": 13,  # M
    "to_date_plan": 15,  # O
    "to_date_actual": 17,  # Q
    "to_date_variance": 18,  # R
}


def _to_percent(value: typing.Any) -> decimal.Decimal | None:
    """Excel stores 21.77% as the fraction 0.2177 — convert to 21.77 for storage."""
    if value is None or not isinstance(value, (int, float, decimal.Decimal)):
        return None
    return decimal.Decimal(str(round(float(value) * 100, 2)))


def _to_date(value: typing.Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


class ProjectScurveCRUDRepository(BaseCRUDRepository):
    async def import_excel(
        self,
        project_id: str,
        file_content: bytes,
        file_name: str,
        uploaded_by_account_id: str | None,
    ) -> dict[str, typing.Any]:
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

        for sheet_name in ("Exe Sum", "S-Overall"):
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan di file Excel.")

        progress_rows = self._parse_exe_sum(workbook["Exe Sum"])
        summary_rows = self._parse_s_overall(workbook["S-Overall"])

        # periode_data ("as of" date) = the latest week whose actual_cumulative is filled in —
        # weeks after that are forecast-only (plan without actual).
        reported_dates = [row["progress_date"] for row in summary_rows if row["actual_cumulative"] is not None]
        if not reported_dates:
            raise ValueError("Tidak ditemukan tanggal laporan (periode_data) yang valid di sheet S-Overall.")
        periode_data = max(reported_dates)

        file_path = self._save_file(project_id=project_id, file_name=file_name, file_content=file_content)

        # project_progress: replace — this sheet only ever describes ONE snapshot (the
        # current report), so old rows for this project are stale the moment a new file lands.
        await self.async_session.execute(
            sqlalchemy.delete(ProjectProgress).where(ProjectProgress.project_id == project_id)
        )
        for row in progress_rows:
            self.async_session.add(
                ProjectProgress(id=uuid.uuid4(), project_id=project_id, periode_data=periode_data, **row)
            )

        # project_progress_summary: upsert by date — each file re-supplies the whole date
        # range and later reports can revise a week's actuals, so dates already stored just
        # get their values refreshed instead of being duplicated.
        existing_stmt = sqlalchemy.select(ProjectProgressSummary).where(
            ProjectProgressSummary.project_id == project_id
        )
        existing_by_date = {
            row.progress_date: row for row in (await self.async_session.execute(existing_stmt)).scalars().all()
        }
        for row in summary_rows:
            existing = existing_by_date.get(row["progress_date"])
            if existing is not None:
                for key, value in row.items():
                    setattr(existing, key, value)
            else:
                self.async_session.add(ProjectProgressSummary(id=uuid.uuid4(), project_id=project_id, **row))

        upload_record = ProjectScurveUpload(
            id=uuid.uuid4(),
            project_id=project_id,
            file_name=file_name,
            file_path=str(file_path),
            uploaded_by_account_id=uploaded_by_account_id,
        )
        self.async_session.add(upload_record)

        await self.async_session.commit()

        return {
            "upload_id": str(upload_record.id),
            "periode_data": periode_data,
            "progress_items": len(progress_rows),
            "summary_weeks": len(summary_rows),
        }

    def _save_file(self, project_id: str, file_name: str, file_content: bytes) -> pathlib.Path:
        project_dir = UPLOAD_ROOT / str(project_id)
        project_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = pathlib.Path(file_name).name  # strip any path components
        destination = project_dir / f"{timestamp}_{safe_name}"
        destination.write_bytes(file_content)

        return destination

    def _parse_exe_sum(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, typing.Any]]:
        """Read Table 1.1 (Discipline Weekly Summary) starting at PROGRESS_FIRST_DATA_ROW.
        Keeps reading rows — skipping blank/spacer ones — until it hits a row whose
        Description is "Overall" (the table's own total row), which marks the end.
        Adding more discipline rows above "Overall" in future templates needs no code
        change, since nothing here assumes a fixed row count.
        """
        rows: list[dict[str, typing.Any]] = []
        row_idx = PROGRESS_FIRST_DATA_ROW
        max_row_idx = PROGRESS_FIRST_DATA_ROW + PROGRESS_MAX_ROWS_SCANNED

        while True:
            if row_idx > max_row_idx:
                raise ValueError(
                    "Tidak ditemukan baris 'Overall' di sheet Exe Sum — pastikan Table 1.1 "
                    "punya baris total 'Overall' sebagai penanda akhir tabel."
                )

            description = sheet.cell(row=row_idx, column=PROGRESS_COLUMNS["description"]).value
            description_str = str(description).strip() if description is not None else ""

            if description_str.strip().lower() == "overall":
                break  # end of Table 1.1 — everything above this is a discipline row

            item_no = sheet.cell(row=row_idx, column=PROGRESS_COLUMNS["item_no"]).value
            if not isinstance(item_no, (int, float)):
                row_idx += 1
                continue  # blank/spacer row — skip, keep scanning for "Overall"

            row: dict[str, typing.Any] = {
                "item_no": int(item_no),
                "description": description_str,
                "wf": _to_percent(sheet.cell(row=row_idx, column=PROGRESS_COLUMNS["wf"]).value),
            }
            for field in (
                "previous_week_plan",
                "previous_week_actual",
                "previous_week_variance",
                "this_week_plan",
                "this_week_actual",
                "this_week_variance",
                "to_date_plan",
                "to_date_actual",
                "to_date_variance",
            ):
                row[field] = _to_percent(sheet.cell(row=row_idx, column=PROGRESS_COLUMNS[field]).value)

            rows.append(row)
            row_idx += 1

        return rows

    def _parse_s_overall(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, typing.Any]]:
        date_by_col: dict[int, datetime.date] = {}
        col = SUMMARY_FIRST_DATA_COL
        while True:
            raw_date = sheet.cell(row=SUMMARY_DATE_ROW, column=col).value
            parsed_date = _to_date(raw_date)
            if parsed_date is None:
                break
            date_by_col[col] = parsed_date
            col += 1

        rows: list[dict[str, typing.Any]] = []
        for col, progress_date in date_by_col.items():
            row = {"progress_date": progress_date}
            for field, metric_row in SUMMARY_METRIC_ROWS.items():
                row[field] = _to_percent(sheet.cell(row=metric_row, column=col).value)
            rows.append(row)

        return rows

    async def read_progress(self, project_id: str) -> typing.Sequence[ProjectProgress]:
        stmt = (
            sqlalchemy.select(ProjectProgress)
            .where(ProjectProgress.project_id == project_id)
            .order_by(ProjectProgress.item_no)
        )
        return (await self.async_session.execute(stmt)).scalars().all()

    async def read_summary(self, project_id: str) -> typing.Sequence[ProjectProgressSummary]:
        stmt = (
            sqlalchemy.select(ProjectProgressSummary)
            .where(ProjectProgressSummary.project_id == project_id)
            .order_by(ProjectProgressSummary.progress_date)
        )
        return (await self.async_session.execute(stmt)).scalars().all()

    async def read_uploads(self, project_id: str) -> typing.Sequence[ProjectScurveUpload]:
        stmt = (
            sqlalchemy.select(ProjectScurveUpload)
            .where(ProjectScurveUpload.project_id == project_id)
            .order_by(ProjectScurveUpload.uploaded_at.desc())
        )
        return (await self.async_session.execute(stmt)).scalars().all()

    async def read_upload_by_id(self, project_id: str, upload_id: str) -> ProjectScurveUpload | None:
        stmt = sqlalchemy.select(ProjectScurveUpload).where(
            ProjectScurveUpload.id == upload_id, ProjectScurveUpload.project_id == project_id
        )
        return (await self.async_session.execute(stmt)).scalar_one_or_none()
